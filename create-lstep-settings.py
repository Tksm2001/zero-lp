from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Color definitions
DARK_BG = "1A1208"
GOLD = "B8860B"
HEADER_FILL = PatternFill("solid", fgColor="2C2010")
GOLD_FILL = PatternFill("solid", fgColor="B8860B")
LIGHT_FILL = PatternFill("solid", fgColor="F5EDD8")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
BLUE_FILL = PatternFill("solid", fgColor="DBEAF7")
GREEN_FILL = PatternFill("solid", fgColor="D5F0D5")
ORANGE_FILL = PatternFill("solid", fgColor="FDE8D0")
RED_FILL = PatternFill("solid", fgColor="F5D5D5")
PURPLE_FILL = PatternFill("solid", fgColor="E8D5F0")
BORDER = Border(
    left=Side(style="thin", color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin", color="999999"),
    bottom=Side(style="thin", color="999999"),
)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, color="B8860B", size=14)
BODY_FONT = Font(name="Arial", size=10, color="333333")
BOLD_FONT = Font(name="Arial", bold=True, size=10, color="333333")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_body(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.alignment = LEFT_WRAP
        cell.border = BORDER
        if fill:
            cell.fill = fill

# ========== Sheet 1: タグ一覧 ==========
ws1 = wb.active
ws1.title = "タグ一覧"
ws1.sheet_properties.tabColor = "B8860B"

ws1.merge_cells("A1:F1")
ws1["A1"] = "Lステップ タグ設計（全16タグ）"
ws1["A1"].font = TITLE_FONT
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

headers1 = ["タグID", "タグ名", "カテゴリ", "付与方法", "付与タイミング", "説明"]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, 6)

tags = [
    [1, "seg_時間", "セグメント", "自動", "Day0 4択「自分の時間がほしい」選択時", "時間不足に悩むセグメント"],
    [2, "seg_収入", "セグメント", "自動", "Day0 4択「収入を増やしたい」選択時", "収入不足に悩むセグメント"],
    [3, "seg_脱出", "セグメント", "自動", "Day0 4択「今の環境から抜け出したい」選択時", "環境変化を望むセグメント"],
    [4, "seg_家族", "セグメント", "自動", "Day0 4択「家族との時間を大切にしたい」選択時", "家族優先セグメント"],
    [5, "rank_A", "ランク", "手動", "Day0-3で2往復以上のやり取り成立", "最優先対応。時間の80%を投下"],
    [6, "rank_B", "ランク", "手動", "Day0-3で1回返信あり", "軽い手動フォロー+自動配信"],
    [7, "rank_C", "ランク", "自動", "Day3終了時に既読のみ（返信なし）", "自動配信のみ"],
    [8, "rank_D", "ランク", "自動", "Day3終了時に未読", "長期育成シナリオへ移行"],
    [9, "route_加速", "ルート", "手動", "Day0返信で3条件中2つ該当", "自動配信停止→手動チャットに完全移行"],
    [10, "route_標準", "ルート", "自動", "加速ルート非該当者に自動付与", "Day0-13自動配信を継続"],
    [11, "cp_共感", "チェックポイント", "手動", "相手が自分の話をし始めた", "CP①：関係構築の第一歩"],
    [12, "cp_痛み", "チェックポイント", "手動", "現状への不満を言語化した", "CP②：変化の動機が明確化"],
    [13, "cp_解放", "チェックポイント", "手動", "「私のせいじゃない」的な発言", "CP③：自己否定からの解放"],
    [14, "cp_投資観", "チェックポイント", "手動+自動", "「お金かけてでも変わりたい」or Day9で「わかる気がする」選択", "CP④：高額商材への心理的準備完了"],
    [15, "cp_信頼移転", "チェックポイント", "手動", "「あいりさんがいうなら」的な発言", "CP⑤：トスアップ可能状態"],
    [16, "migration_個人LINE", "移行", "自動", "Day10で個人LINE追加確認時", "個人LINEへの移行完了"],
]

cat_fills = {
    "セグメント": BLUE_FILL,
    "ランク": GREEN_FILL,
    "ルート": ORANGE_FILL,
    "チェックポイント": PURPLE_FILL,
    "移行": LIGHT_FILL,
}

for i, tag in enumerate(tags):
    row = i + 4
    for j, val in enumerate(tag):
        ws1.cell(row=row, column=j + 1, value=val)
    fill = cat_fills.get(tag[2], WHITE_FILL)
    style_body(ws1, row, 6, fill)
    ws1.cell(row=row, column=1).alignment = CENTER

ws1.column_dimensions["A"].width = 8
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 16
ws1.column_dimensions["D"].width = 12
ws1.column_dimensions["E"].width = 50
ws1.column_dimensions["F"].width = 35

# ========== Sheet 2: シナリオ一覧 ==========
ws2 = wb.create_sheet("シナリオ一覧")
ws2.sheet_properties.tabColor = "B8860B"

ws2.merge_cells("A1:G1")
ws2["A1"] = "Lステップ シナリオ設計（全6シナリオ）"
ws2["A1"].font = TITLE_FONT

headers2 = ["ID", "シナリオ名", "対象者条件", "開始トリガー", "期間", "総配信数", "概要"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 7)

scenarios = [
    [1, "Day0_制圧戦", "全登録者", "友だち追加", "0〜3時間", "4通(+分岐4)", "登録直後のあいさつ・PDF配布・欲求4択・思い出す問い・リマインド"],
    [2, "Day1-3_関係構築", "route_標準 AND NOT rank_D", "Day1 08:04", "3日間", "5通", "音声メッセージ・体験談(4分岐)・動画・恐怖の2択"],
    [3, "Day4-9_価値観書き換え", "route_標準 AND NOT rank_D", "Day4 21:02", "6日間", "7通(+分岐4)", "解放宣言・行動促進・期限意識・師匠伏線・比較ストーリー・温度確認"],
    [4, "Day10-13_トスアップ", "route_標準 AND NOT rank_D", "Day10 21:02", "4日間", "3通+空白1日", "個人LINE移行・師匠公開・空白日・逆説的クロージング"],
    [5, "加速ルート_手動支援", "route_加速", "加速判定時", "最短3日", "手動（自動停止）", "CP①-⑤を会話内で通過。担当者フルチャット対応"],
    [6, "長期育成", "rank_C OR rank_D, アポ未確定", "Day14以降", "12ヶ月", "24通(月2回)", "成功者ストーリー・マインド・数字・近況の4パターンローテ"],
]

for i, s in enumerate(scenarios):
    row = i + 4
    for j, val in enumerate(s):
        ws2.cell(row=row, column=j + 1, value=val)
    style_body(ws2, row, 7)
    ws2.cell(row=row, column=1).alignment = CENTER

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 25
ws2.column_dimensions["C"].width = 30
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 18
ws2.column_dimensions["G"].width = 55

# ========== Sheet 3: 配信スケジュール ==========
ws3 = wb.create_sheet("配信スケジュール")
ws3.sheet_properties.tabColor = "B8860B"

ws3.merge_cells("A1:J1")
ws3["A1"] = "全配信スケジュール（Day0〜13 + 長期育成）"
ws3["A1"].font = TITLE_FONT

headers3 = ["Day", "配信時間", "通番号", "メッセージ概要", "形式", "文字数", "セグメント分岐", "タグ条件", "自動/手動", "備考"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, 10)

schedule = [
    ["Day0", "即時(0分)", "D0-1", "あいさつ＋PDF配布", "テキスト+PDF", "200字(中)", "なし", "全員", "自動", "冒頭1行がプッシュ通知に表示"],
    ["Day0", "+30秒", "D0-2", "欲求4択クイックリプライ", "テキスト+QR", "100字(軽)", "なし", "全員", "自動", "4択でセグメントタグ付与"],
    ["Day0", "即時(4択後)", "D0-3a", "思い出す問い（時間）", "テキスト", "250字(中)", "seg_時間", "seg_時間", "自動", "「いつまで続くんだろう」"],
    ["Day0", "即時(4択後)", "D0-3b", "思い出す問い（収入）", "テキスト", "250字(中)", "seg_収入", "seg_収入", "自動", "「眠れなかった夜ある？」"],
    ["Day0", "即時(4択後)", "D0-3c", "思い出す問い（脱出）", "テキスト", "250字(中)", "seg_脱出", "seg_脱出", "自動", "「辞めたいと思った瞬間」"],
    ["Day0", "即時(4択後)", "D0-3d", "思い出す問い（家族）", "テキスト", "250字(中)", "seg_家族", "seg_家族", "自動", "「ごめんねって思った瞬間」"],
    ["Day0", "+3時間", "D0-4", "PDFリマインド", "テキスト", "80字(軽)", "なし", "通3未返信者のみ", "自動", "「5ページ目だけ見てね」"],
    ["Day1", "08:04", "D1-1a", "あいりの音声メッセージ", "音声30秒", "-", "なし", "route_標準", "自動", "朝の落ち着いたトーン"],
    ["Day1", "08:04", "D1-1b", "音声補足テキスト", "テキスト", "60字(軽)", "なし", "route_標準", "自動", "「夜に続き話すね」"],
    ["Day1", "21:02", "D1-2a", "体験談＋問い（時間）", "テキスト", "400字(重)", "seg_時間", "route_標準 & seg_時間", "自動", "「戻りたくない朝」+分岐問い"],
    ["Day1", "21:02", "D1-2b", "体験談＋問い（収入）", "テキスト", "400字(重)", "seg_収入", "route_標準 & seg_収入", "自動", "「戻りたくない朝」+分岐問い"],
    ["Day1", "21:02", "D1-2c", "体験談＋問い（脱出）", "テキスト", "400字(重)", "seg_脱出", "route_標準 & seg_脱出", "自動", "「戻りたくない朝」+分岐問い"],
    ["Day1", "21:02", "D1-2d", "体験談＋問い（家族）", "テキスト", "400字(重)", "seg_家族", "route_標準 & seg_家族", "自動", "「戻りたくない朝」+分岐問い"],
    ["Day2", "21:03", "D2-1a", "動画導入テキスト", "テキスト", "120字(軽)", "なし", "route_標準", "自動", "「顔見て話したい」"],
    ["Day2", "21:03", "D2-1b", "動画リッチメッセージ", "動画60秒", "-", "なし", "route_標準", "自動", "あいり顔出し。自動再生推奨"],
    ["Day2", "21:03", "D2-1c", "恐怖の2択テキスト", "テキスト", "150字(中)", "なし", "route_標準", "自動", "「3年後変わらない vs 失敗」"],
    ["Day3", "08:04", "D3-1", "恐怖の2択クイックリプライ", "テキスト+QR", "250字(中)", "なし", "route_標準", "自動", "2択：変わらない方が怖い/まだ怖い"],
    ["Day3", "即時", "D3-1f", "「まだ怖い」選択者フォロー", "テキスト", "80字(軽)", "なし", "「まだ怖い」選択者", "自動", "「怖いのが普通」で安心させる"],
    ["Day4", "21:02", "D4-1", "「あなたのせいじゃない」解放宣言", "テキスト", "500字(重)", "なし", "route_標準 NOT rank_D", "自動", "CP③解放。地図のメタファー"],
    ["Day5", "21:02", "D5-1", "「考える前に動く」+OQ", "テキスト", "350字(中)", "なし", "route_標準 NOT rank_D", "自動", "★手動介入①の対象通"],
    ["Day6", "21:03", "D6-1a", "「あと何年？」（時間）", "テキスト", "300字(中)", "seg_時間", "route_標準 & seg_時間", "自動", "「逃げ道がほしかった」"],
    ["Day6", "21:03", "D6-1b", "「あと何年？」（収入）", "テキスト", "300字(中)", "seg_収入", "route_標準 & seg_収入", "自動", "「もうひとつの収入」"],
    ["Day6", "21:03", "D6-1c", "「あと何年？」（脱出）", "テキスト", "300字(中)", "seg_脱出", "route_標準 & seg_脱出", "自動", "「辞めても大丈夫な自分」"],
    ["Day6", "21:03", "D6-1d", "「あと何年？」（家族）", "テキスト", "300字(中)", "seg_家族", "route_標準 & seg_家族", "自動", "「子どもの成長は待ってくれない」"],
    ["Day7", "21:02", "D7-1", "「最初の1円」+師匠セリフ初出", "テキスト", "600字(重)", "なし", "route_標準 NOT rank_D", "自動", "★師匠「才能じゃない。環境と地図」"],
    ["Day8", "21:03", "D8-1", "「Aちゃんの話」独学vs.サポート", "テキスト", "500字(重)", "なし", "route_標準 NOT rank_D", "自動", "独学のリスクを物語で体感"],
    ["Day9", "21:02", "D9-1", "投資観インストール+温度確認QR", "テキスト+QR", "400字(重)", "なし", "route_標準 NOT rank_D", "自動", "2択：わかる気がする/まだ怖い"],
    ["Day9", "即時", "D9-1a", "「わかる気がする」返答", "テキスト", "100字(軽)", "なし", "「わかる」選択者", "自動", "cp_投資観タグ付与+Day10伏線"],
    ["Day9", "即時", "D9-1b", "「まだ怖い」返答", "テキスト", "100字(軽)", "なし", "「怖い」選択者", "自動", "「あなたのペースで」"],
    ["Day10", "21:02", "D10-1", "個人LINE誘導", "テキスト+URL", "350字(中)", "なし", "route_標準 NOT rank_D", "自動", "個人LINE URL/QR送付"],
    ["Day11", "21:02", "D11-1", "師匠の正体を明かす", "テキスト", "450字(重)", "なし", "migration_個人LINE", "手動推奨", "★伏線回収。「明日は送らない」予告"],
    ["Day12", "—", "—", "空白日（個人LINE）", "—", "—", "—", "migration_個人LINE", "—", "意図的な不在感。飢餓感の創出"],
    ["Day12", "21:02", "D12-1", "軽いフォロー（公式LINE）", "テキスト", "80字(軽)", "なし", "NOT migration_個人LINE", "自動", "未移行者への最後のリマインド"],
    ["Day13", "21:02", "D13-1", "逆説的CL+アポ打診", "テキスト", "500字(重)", "なし", "migration_個人LINE", "手動", "★「無理にとは言わない」+30分電話"],
    ["Day13", "即時", "D13-2", "アポ日程調整", "テキスト", "100字(軽)", "なし", "「話したい」返信者", "手動", "3候補日を提示"],
    ["Day13", "即時", "D13-3", "アポ確定フォロー", "テキスト", "80字(軽)", "なし", "アポ確定者", "手動", "★手動介入③「緊張してる」"],
    ["Day15", "+48h", "D13-F", "未返信者フォロー", "テキスト", "60字(軽)", "なし", "Day13未返信者", "手動", "「いつでもここにいるから」→長期育成へ"],
]

day_fills = {
    "Day0": PatternFill("solid", fgColor="FFF8E1"),
    "Day1": PatternFill("solid", fgColor="FFF3E0"),
    "Day2": PatternFill("solid", fgColor="FFF3E0"),
    "Day3": PatternFill("solid", fgColor="FFF3E0"),
    "Day4": PatternFill("solid", fgColor="E8F5E9"),
    "Day5": PatternFill("solid", fgColor="E8F5E9"),
    "Day6": PatternFill("solid", fgColor="E8F5E9"),
    "Day7": PatternFill("solid", fgColor="E8F5E9"),
    "Day8": PatternFill("solid", fgColor="E8F5E9"),
    "Day9": PatternFill("solid", fgColor="E8F5E9"),
    "Day10": PatternFill("solid", fgColor="E3F2FD"),
    "Day11": PatternFill("solid", fgColor="E3F2FD"),
    "Day12": PatternFill("solid", fgColor="E3F2FD"),
    "Day13": PatternFill("solid", fgColor="E3F2FD"),
    "Day15": PatternFill("solid", fgColor="F3E5F5"),
}

for i, s in enumerate(schedule):
    row = i + 4
    for j, val in enumerate(s):
        ws3.cell(row=row, column=j + 1, value=val)
    fill = day_fills.get(s[0], WHITE_FILL)
    style_body(ws3, row, 10, fill)
    ws3.cell(row=row, column=1).alignment = CENTER
    ws3.cell(row=row, column=9).alignment = CENTER

ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 32
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 12
ws3.column_dimensions["G"].width = 14
ws3.column_dimensions["H"].width = 28
ws3.column_dimensions["I"].width = 12
ws3.column_dimensions["J"].width = 35

# ========== Sheet 4: 手動介入ポイント ==========
ws4 = wb.create_sheet("手動介入ポイント")
ws4.sheet_properties.tabColor = "A83232"

ws4.merge_cells("A1:G1")
ws4["A1"] = "手動介入ポイント（3タイミング）"
ws4["A1"].font = TITLE_FONT

headers4 = ["タイミング", "Day", "対象", "アクション", "所要時間", "テンプレ例", "月間工数"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, 7)

interventions = [
    ["①Day5 返信フォロー", "Day5-6", "Day5のOQに返信した人", "個別返信：相手の言葉引用→共感→可能性提示", "1人3分", "「〇〇したいんだね。私も最初そう思ってた。実はそれ、意外と遠くないよ」", "20人×3分＝60分/月"],
    ["②Day9 前向き者フォロー", "Day9-10", "「わかる気がする」選択者", "手動で一言追加送信", "1人2分", "「実はあなたに話したいことがある」", "30人×2分＝60分/月"],
    ["③Day13 アポ後フォロー", "Day13以降", "アポ確定者", "確定直後に手動送信。キャンセル防止", "1人3分", "「楽しみにしてるね。少し緊張してる」", "15人×3分＝45分/月"],
]

for i, inv in enumerate(interventions):
    row = i + 4
    for j, val in enumerate(inv):
        ws4.cell(row=row, column=j + 1, value=val)
    style_body(ws4, row, 7, ORANGE_FILL)

ws4.merge_cells("A8:G8")
ws4["A8"] = "合計月間工数：約165分（2時間45分）"
ws4["A8"].font = Font(name="Arial", bold=True, size=12, color="A83232")

ws4.merge_cells("A10:G10")
ws4["A10"] = "担当者1日のルーティン（合計90分）"
ws4["A10"].font = Font(name="Arial", bold=True, size=12, color="B8860B")

routine_headers = ["時間帯", "時間", "内容", "対象", "ポイント", "", ""]
for i, h in enumerate(routine_headers, 1):
    ws4.cell(row=11, column=i, value=h)
style_header(ws4, 11, 7)

routines = [
    ["朝 8:00-8:30", "30分", "前夜の返信チェック＋Aランク個別返信作成", "Aランク＋加速ルート", "相手の言葉を引用して返信開始"],
    ["昼 12:00-12:15", "15分", "加速ルートのみ対応", "route_加速", "短文で即レス。会話のリズムを維持"],
    ["夜 21:00-21:45", "45分", "当日の返信チェック＋加速ルート集中チャット", "Aランク＋加速ルート", "1日の中で最も深い会話をする時間帯"],
]

for i, r in enumerate(routines):
    row = i + 12
    for j, val in enumerate(r):
        ws4.cell(row=row, column=j + 1, value=val)
    style_body(ws4, row, 5, LIGHT_FILL)

ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 12
ws4.column_dimensions["C"].width = 40
ws4.column_dimensions["D"].width = 25
ws4.column_dimensions["E"].width = 40
ws4.column_dimensions["F"].width = 20
ws4.column_dimensions["G"].width = 22

# ========== Sheet 5: KPIトラッカー ==========
ws5 = wb.create_sheet("KPIトラッカー")
ws5.sheet_properties.tabColor = "2E6B8A"

ws5.merge_cells("A1:F1")
ws5["A1"] = "月次KPIトラッカー"
ws5["A1"].font = TITLE_FONT

headers5 = ["指標名", "目標値", "月1実績", "月2実績", "月3実績", "備考"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, 6)

kpis = [
    ["LP流入数", "300人/月", "", "", "", "広告・SNS・紹介合計"],
    ["公式LINE登録数", "100人/月 (33%)", "", "", "", "LP→LINE登録率"],
    ["Day0 4択回答率", "70%", "", "", "", "クイックリプライ応答率"],
    ["Day0 思い出す問い返信率", "30%", "", "", "", "自由返信率"],
    ["加速ルート判定数", "10-15人 (10-15%)", "", "", "", "3条件中2つ該当者"],
    ["Day3 Aランク確定数", "15人 (15%)", "", "", "", "2往復以上のやり取り成立"],
    ["Day5 OQ返信数", "20人 (20%)", "", "", "", "オープンクエスチョン返信者"],
    ["Day9「わかる気がする」数", "10人 (10%)", "", "", "", "温度確認QR応答"],
    ["Day10 個人LINE追加数", "7-8人 (7-8%)", "", "", "", "公式→個人移行"],
    ["Day13 アポ希望数", "5人 (5%)", "", "", "", "「話したい」返信者"],
    ["3回アポ完走数", "3-4人 (3-4%)", "", "", "", "3回の電話アポ完了"],
    ["成約数", "1-2人 (1-2%)", "", "", "", "50〜200万円の契約"],
    ["月間売上", "50-400万円", "", "", "", "成約数×商材単価"],
    ["加速ルートアポ到達率", "50-60%", "", "", "", "加速ルート→アポ"],
    ["標準ルートアポ到達率", "5-10%", "", "", "", "標準ルート→アポ"],
    ["長期育成復活率", "5-10%/年", "", "", "", "12ヶ月間での返信復活"],
]

for i, kpi in enumerate(kpis):
    row = i + 4
    for j, val in enumerate(kpi):
        ws5.cell(row=row, column=j + 1, value=val)
    fill = LIGHT_FILL if i % 2 == 0 else WHITE_FILL
    style_body(ws5, row, 6, fill)

# Conversion rate formulas
ws5.cell(row=21, column=1, value="")
ws5.merge_cells("A22:F22")
ws5["A22"] = "ファネル全体CVR（登録→成約）"
ws5["A22"].font = Font(name="Arial", bold=True, size=12, color="2E6B8A")

ws5.column_dimensions["A"].width = 30
ws5.column_dimensions["B"].width = 20
ws5.column_dimensions["C"].width = 15
ws5.column_dimensions["D"].width = 15
ws5.column_dimensions["E"].width = 15
ws5.column_dimensions["F"].width = 30

# ========== Sheet 6: メッセージ本文一覧 ==========
ws6 = wb.create_sheet("メッセージ本文")
ws6.sheet_properties.tabColor = "C47D1A"

ws6.merge_cells("A1:D1")
ws6["A1"] = "全メッセージ本文（Lステップ転記用）"
ws6["A1"].font = TITLE_FONT

headers6 = ["通番号", "メッセージ概要", "本文", "注意事項"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=3, column=i, value=h)
style_header(ws6, 3, 4)

messages = [
    ["D0-1", "あいさつ＋PDF配布", "登録ありがとうございます！\nあいりです☺️\n\n約束のPDF、すぐ送りますね。\n\n▼ 副業ロードマップ（全12ページ）\n[PDFリンク]\n\n全部読まなくて大丈夫です。\nまず5ページ目だけ見てみてください。\n\n「あ、これ私のことだ」って\n思う人が多いページです。\n\n見たらまた連絡くださいね。\nひとつだけ聞きたいことがあるので🙏", "PDFリンクは実際のURLに差し替え"],
    ["D0-2", "欲求4択", "あ、その前にひとつだけ！\n\n今いちばん「変えたい」って\n思ってることってどれに近い？👇\n\n[クイックリプライ4択]", "QR: ⏰自分の時間がほしい / 💰収入を増やしたい / 🚪今の環境から抜け出したい / 👨‍👩‍👧家族との時間を大切にしたい"],
    ["D0-3a", "思い出す問い（時間）", "ああ、時間かぁ…。\nわかる。私もそうだった。\n\nちょっと聞いてもいい？\n\n「いつまでこの生活続くんだろう」\nって思った瞬間ってある？\n\nたとえば\n\n・朝の満員電車でぼーっとしてる時とか\n・子ども寝かしつけた後にPC開く瞬間とか\n・日曜の夜、明日のこと考えた時とか\n\nなんでもいいよ。\n思い出すだけでいいから。", "seg_時間タグ付与済の人に自動送信"],
    ["D0-3b", "思い出す問い（収入）", "うん、お金のことって\nなかなか人に言えないよね。\n\nちょっと聞いてもいい？\n\nお金のことで\n眠れなかった夜ってある？\n\nたとえば\n\n・給料日に通帳見て「…これだけ？」って時とか\n・クレカの引き落とし日が近づいてくる時とか\n・子どもの習い事の月謝を考える時とか\n\nなんでもいいよ。\n思い出すだけでいいから。", "seg_収入タグ付与済の人に自動送信"],
    ["D0-3c", "思い出す問い（脱出）", "わかる…。\n逃げたいって思うの、全然おかしくないよ。\n\nちょっと聞いてもいい？\n\n「もう辞めたい」って\nいちばん強く思った瞬間ってある？\n\nたとえば\n\n・上司に何か言われた時とか\n・日曜の夜、「また明日か…」って時とか\n・朝のアラームが鳴って動けなかった時とか\n\nなんでもいいよ。\n思い出すだけでいいから。", "seg_脱出タグ付与済の人に自動送信"],
    ["D0-3d", "思い出す問い（家族）", "家族のこと考えてるんだね。\nそれだけで、もう十分すごいと思う。\n\nちょっと聞いてもいい？\n\n子どもに「ごめんね」って\n思った瞬間ってある？\n\nたとえば\n\n・お迎えに間に合わなかった時とか\n・「遊ぼう」って言われて断った時とか\n・寝顔しか見れない日が続いた時とか\n\nなんでもいいよ。\n思い出すだけでいいから。", "seg_家族タグ付与済の人に自動送信"],
    ["D0-4", "PDFリマインド", "さっきのPDF、保存できた？📄\n\nもし忙しかったら、\n5ページ目だけでいいので\nあとで見てみてね。\n\n3分で読めるから🙏", "通3に返信なしの人のみ配信"],
    ["D1-1", "音声メッセージ＋補足", "[音声メッセージ30秒]\n\n音声聴いてくれた？🎧\n\n夜にちゃんと続き話すね。\n楽しみにしててほしい☺️", "音声台本は別途。録音ディレクション参照"],
    ["D1-2a", "体験談＋問い（時間）", "朝の続きね。\n\n私がこの仕事を続けられてる理由。\nそれは「戻りたくない朝」があるから。\n\n2年前、毎朝アラームが鳴るたびに\n天井見て「あと何年これ続けるんだろう」\nって思ってた。\n\n上司は悪い人じゃなかった。\n給料も普通だった。\nでも、心がずっと死んでた。\n\nあなたも昨日\n「いつまで続くんだろう」って\n話してくれたよね。\n\nあの感覚、私もずっとあった。\n\nちょっと聞いていい？\n\n初めて「今の生活変えたい」って\n検索した日のこと、覚えてる？\n\n何がきっかけだった？", "共通パート＋時間セグメント分岐"],
    ["D1-2b", "体験談＋問い（収入）", "朝の続きね。\n\n私がこの仕事を続けられてる理由。\nそれは「戻りたくない朝」があるから。\n\n2年前、毎朝アラームが鳴るたびに\n天井見て「あと何年これ続けるんだろう」\nって思ってた。\n\n上司は悪い人じゃなかった。\n給料も普通だった。\nでも、心がずっと死んでた。\n\nあなたも昨日\nお金のことで眠れない夜が\nあるって話してくれたよね。\n\nあの感覚、私もずっとあった。\n\nちょっと聞いていい？\n\n初めて「副業」とか「稼ぎ方」って\n検索した日のこと、覚えてる？\n\n何がきっかけだった？", "共通パート＋収入セグメント分岐"],
    ["D1-2c", "体験談＋問い（脱出）", "朝の続きね。\n\n私がこの仕事を続けられてる理由。\nそれは「戻りたくない朝」があるから。\n\n2年前、毎朝アラームが鳴るたびに\n天井見て「あと何年これ続けるんだろう」\nって思ってた。\n\n上司は悪い人じゃなかった。\n給料も普通だった。\nでも、心がずっと死んでた。\n\nあなたも昨日\n「辞めたい」って\n強く思った瞬間の話をしてくれたよね。\n\nあの感覚、私もずっとあった。\n\nちょっと聞いていい？\n\nその気持ち、\n誰かに相談したことある？\n…それとも、言えなかった？", "共通パート＋脱出セグメント分岐"],
    ["D1-2d", "体験談＋問い（家族）", "朝の続きね。\n\n私がこの仕事を続けられてる理由。\nそれは「戻りたくない朝」があるから。\n\n2年前、毎朝アラームが鳴るたびに\n天井見て「あと何年これ続けるんだろう」\nって思ってた。\n\n上司は悪い人じゃなかった。\n給料も普通だった。\nでも、心がずっと死んでた。\n\nあなたも昨日\n子どもに「ごめんね」って\n思った瞬間の話をしてくれたよね。\n\nあの気持ち、私もわかる。\n\nちょっと聞いていい？\n\n家族のために何か変えたいって\n初めて思った日のこと、覚えてる？\n\n何がきっかけだった？", "共通パート＋家族セグメント分岐"],
    ["D2-1", "動画＋恐怖の2択", "[動画導入]\n今日は文章じゃなくて\n顔見て話したいなと思って。\n1分だけ。見てほしい。\n\n[動画リッチメッセージ60秒]\n\n[動画後テキスト]\n動画見てくれた？\n\n3年後も今と変わってない自分。\nそれとも、3年後に失敗してる自分。\n\nどっちが怖い？\n\n…正直、私はどっちも怖かった。\nでも「変わらないこと」の方がもっと怖かった。\n\n明日、その話の続きするね。", "動画台本は別途。撮影ディレクション参照"],
    ["D3-1", "恐怖の2択QR", "おはよう☺️\n\n昨日の動画見てくれた？\n\n「3年後も変わってない自分」と\n「3年後に失敗してる自分」\n\nどっちが怖いかって聞いたけど、\n実はね、もうひとつあって。\n\nいちばん怖いのは\n「あの時やっておけば」って\n3年後に後悔してる自分。\n\n変わらないことも、失敗も、\n実はどうにでもなる。\n\nでも「やらなかった後悔」だけは\n取り返しがつかない。\n\n私はそれがいちばん怖かった。\n\nあなたはどう思う？👇\n\n[QR: 😰変わらないことの方が怖い / 🤔まだちょっと怖い]", "「まだ怖い」選択者にフォロー自動送信"],
    ["D4-1", "解放宣言", "今日はちょっと大事な話。\n\nここ数日やり取りしてきて、\nあなたが本気で「変わりたい」って\n思ってるのは伝わってる。\n\nだからこそ、\nひとつだけ先に言わせて。\n\nうまくいかなかったの、\nあなたのせいじゃないよ。\n\n独学で調べて、YouTube見て、本読んで、\nそれでもうまくいかなかった。\nそれは当たり前。\n\nだって「正しい順番」と「正しい環境」を\n誰も教えてくれなかっただけだから。\n\n才能がないわけじゃない。\n努力が足りないわけでもない。\n地図がなかっただけ。\n\n知らない土地をGoogle Mapなしで歩いてたようなもの。\n迷って当然だし、疲れて当然。\n\nだから、自分を責めないで。\nここまで読んでくれてるあなたは\nもう十分がんばってる。\n\n明日は「考える前に動く」って\nどういうことか話すね。\nちょっとだけ、面白い話がある☺️", "CP③解放の到達を狙う通。返信促さない"],
    ["D5-1", "行動の種＋OQ", "昨日の話、読んでくれた？\n\n「あなたのせいじゃない」って言ったけど、\nじゃあ何が原因だったのか。\n\n答えはシンプルで、\n「考えすぎて動けなかった」こと。\n\nこれ、私もそうだった。\n\n「失敗したらどうしよう」\n「自分にできるのかな」\n「もうちょっと調べてから…」\n\nって、ずっと考えてた。\n\nでもある日気づいたんだよね。\n考えてる間は、何も変わらない。\n\nだから今日はひとつだけ聞いていい？\n\nもし今、何のリスクもなくて\n絶対に失敗しないとしたら、\nいちばん最初に何をする？\n\n仕事辞める？引っ越す？新しいこと始める？\n\nなんでもいい。正解はないから。\n思いついたこと、教えてくれたら嬉しい☺️", "★手動介入①対象。返信者に翌日中に個別フォロー"],
    ["D7-1", "最初の1円＋師匠セリフ", "今日はちょっと長くなるけど、大事な話させて。\n\n私が「人生変わった」って思った瞬間の話。\n\n副業始めて最初の3ヶ月、1円も稼げなかった。\nYouTube見て、本読んで、自分なりにやってみたけど、\n何が正しいのかわからなくて。\n\n「やっぱり自分には無理なのかな」って本気で思ってた。\n\nそんな時、ある人に出会った。\nその人に最初に言われたのが、\n\n「才能の問題じゃない。環境と地図の問題だ」\n\nって言葉だった。\n\n正直、最初は「何言ってんだろ」って思った。笑\n\nでもその人の言う通りにやったら、\n2週間後に初めて500円が振り込まれた。\n\n…たった500円。\nでもね、あの振込通知を見た時、手が震えた。\n\n「自分でお金を作れた」\nその感覚は、給料日に振り込まれる25万円より\nずっとずっと大きかった。\n\n500円から始まって、翌月は3万、その次は10万、\n半年後には本業の収入を超えた。\n\n全部、あの人の一言から始まった。\n\n「才能の問題じゃない。環境と地図の問題だ」\n\nこの言葉、覚えておいて。\nあとで意味がわかるから。", "★師匠セリフ初出。Day11で伏線回収"],
    ["D9-1", "投資観＋温度確認", "9日間、読んでくれてありがとう。\n\nここまで付き合ってくれてるあなたに正直に話すね。\n\n私が変われた理由、才能でも努力でもなくて、\n「正しい環境にお金を使った」から。\n\n最初は怖かった。「騙されたらどうしよう」って。\n「そんなお金ないし」って。\n\nでもある時こう考えた。\n\n「今の自分のまま5年過ごすのと、\nここで投資して1年で変わるの、\nどっちが本当に高い？」\n\n答えは明らかだった。\n\n投資って言っても、株とか不動産の話じゃない。\n「未来の自分を前倒しにする」ってこと。\n\n1年後になれる自分に、3ヶ月でなれるとしたら。\nその9ヶ月の差に価値を感じるかどうか。\n\nここまで読んで、正直どう思った？👇\n\n[QR: 💡わかる気がする / 😟まだ怖い]", "★手動介入②対象。「わかる」選択者にDay10伏線"],
    ["D10-1", "個人LINE誘導", "10日間、ありがとう。\n\n今日はひとつお願いがあって。\n\nここから先の話、公式LINEではちょっとしにくくて。\n\n理由はシンプルで、\n公式LINEって他の人にも同じ配信が届く仕組みだから。\n\nでも、ここから先はあなただけに話したいことがある。\n\nだから、私の個人のLINEを送るね。\n\n[個人LINEのURL/QR]\n\n追加してくれたら、そっちで続き話すね。\n\nもちろん、追加するかどうかはあなた次第。\n無理にとは言わないし、\nこのまま公式LINEで見てくれるだけでも全然いい。\n\nでも、本気で変わりたいなら、\nここから先の話は聞いてほしい。\n\n待ってるね☺️", "個人LINE URLは実際のものに差し替え"],
    ["D13-1", "逆説的CL＋アポ打診", "昨日は1日考えてくれた？\n考えてくれただけでありがとうって思う。\n\n今日で13日目。\n最初にLINE登録してくれた時からずっと見てきたけど、\nあなたは本気だと思ってる。\n\nだからこそ正直に言うね。\n\n私にできることは、ここまで。\nこれ以上はLINEのメッセージじゃ伝えられない。\n\n私が変われたのは、\n【師匠の名前】さんに直接話を聞いてもらえたから。\n\nLINEで読むのと、\n声を聞きながら自分の状況を話すのは、まったく別物だから。\n\nだから、もし「話を聞いてみたい」って思ってくれるなら、\n30分だけ、電話で話せる時間を作ってほしい。\n\nあなたの話を聞いて、\nあなたに合ったやり方があるか一緒に考えたい。\n\nただし、ひとつだけ。\n無理にとは絶対に言わない。\n焦る必要もない。「今じゃない」と思ったら、それでいい。\n\nでも、「本気で変わりたい」なら、\n私は責任を持って紹介する。\n\n自分のペースで決めてね。\n話してみたいと思ったら「話したい」って返信してくれるだけでいい。\n日程はこっちで調整するから☺️", "★手動送信。【師匠の名前】は実名に差し替え"],
]

for i, msg in enumerate(messages):
    row = i + 4
    for j, val in enumerate(msg):
        ws6.cell(row=row, column=j + 1, value=val)
    fill = LIGHT_FILL if i % 2 == 0 else WHITE_FILL
    style_body(ws6, row, 4, fill)

ws6.column_dimensions["A"].width = 10
ws6.column_dimensions["B"].width = 25
ws6.column_dimensions["C"].width = 80
ws6.column_dimensions["D"].width = 35

for row in ws6.iter_rows(min_row=4, max_row=len(messages) + 3, min_col=3, max_col=3):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws6.row_dimensions[cell.row].height = 200

output_path = "/Users/kt/Documents/zero-lp/lstep-settings.xlsx"
wb.save(output_path)
print(f"DONE: {output_path}")
